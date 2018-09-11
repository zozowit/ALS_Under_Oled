import openpyxl
import os

def fetch_data(file_path, integrate):
    print('fetch data from:' + file_path)
    
    [xlsx_dir, xlsx_extra] = file_path.split('.', 1)

    wb = openpyxl.load_workbook(file_path)

    if integrate is False:
        txt_name = xlsx_dir + '.txt'
    else:
        txt_name = os.path.join(os.getcwd(), 'data.txt')
        
    print(txt_name)
    f_txt = open(txt_name, 'a+')

    ws = wb[wb.sheetnames[0]]
    
    for row in ws.iter_rows():
        temps = ''
        for cell in row:
            temps = temps + str(cell.value) + ','

        f_txt.write(temps[:-1] + '\n')

    f_txt.close()

def search_path():
    for root, dirs, files in os.walk('.', topdown=False):
        for name in files:
            f_path = os.path.join(os.getcwd(), name)
            print(f_path)
            (f_name, f_ext) = name.split('.', 1)
            if (f_ext == 'xlsx'):
                fetch_data(f_path, True)

def main():
    xlsx_file = input('input the xlsx location or input X to integrate all the file in current patch:')
    
    if xlsx_file == 'X':
        search_path()
    else:
        fetch_data(xlsx_file, False)

if __name__ == "__main__":
    main()
